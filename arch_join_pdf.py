import os
import openpyxl
from PyPDF2 import PdfMerger

# Файл создал Кирилл Смирнов: Архитектор 1 категории А101
# По вопросам доработки пишите в телеграмм @s_kirill94

# Путь к файлу и листам
template_file = 'template.xlsx'
files_sheet_name = 'files'
files_sort_sheet_name = 'files_sort'
filename = 'filename'

# 1. Получаем список всех PDF файлов в текущей папке
pdf_files = [f for f in os.listdir() if f.endswith('.pdf')]

# 2. Открываем Excel файл и записываем имена PDF файлов в лист 'files'
wb = openpyxl.load_workbook(template_file)
ws_files = wb[files_sheet_name]

# Очищаем лист перед записью новых данных
ws_files.delete_rows(1, ws_files.max_row)

for index, pdf_file in enumerate(pdf_files, start=1):
    ws_files.cell(row=index, column=1, value=pdf_file)

wb.save(template_file)

# 3. Читаем имена файлов с листа 'files_sort'
ws_sort = wb[files_sort_sheet_name]
files_to_merge = [cell.value for cell in ws_sort['A'] if cell.value is not None]

# 4. Читаем имя с итогового файла 'filename'
ws_name = wb[filename]
filename = ws_name["A1"].value

# 5. Объединяем PDF файлы
merger = PdfMerger()

for pdf_name in files_to_merge:
    if pdf_name in pdf_files:
        merger.append(pdf_name)
    else:
        print(f"Файл не найден в текущей папке.")

directory_up = os.path.abspath(os.path.join(os.getcwd(), os.path.pardir))

merged_pdf_filename = str(filename) + ".pdf"
full_name_pdf = os.path.join(directory_up, merged_pdf_filename)
merger.write(full_name_pdf)
merger.close()
os.startfile(full_name_pdf)


print("PDF файлы объединены и сохранены папкой выше")
